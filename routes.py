from app import app, login_manager
from flask import render_template, redirect, request, url_for, send_file, flash
from models import db, User, AuthUser, Vacacion
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.security import check_password_hash
from sqlalchemy import func
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
import requests
import os
import io

# =========================
# LOGIN MANAGER
# =========================
@login_manager.user_loader
def load_user(user_id):
    return AuthUser.query.get(int(user_id))


# =========================
# LOGIN
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():

    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        user = AuthUser.query.filter_by(username=username).first()

        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for("loading"))

        return render_template(
            "login.html",
            error="ACCESS DENIED – INVALID CREDENTIALS"
        )

    return render_template("login.html")


# =========================
# LOADING
# =========================
@app.route("/loading")
@login_required
def loading():
    return render_template("loading.html")


# =========================
# LOGOUT
# =========================
@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# =========================
# DASHBOARD
# =========================
@app.route("/")
@login_required
def dashboard():

    users = User.query.all()
    total_users = len(users)

    activos = User.query.filter_by(activo=True).count()
    inactivos = User.query.filter_by(activo=False).count()

    equipos = db.session.query(
        User.equipo,
        func.count(User.id)
    ).group_by(User.equipo).all()

    return render_template(
        "dashboard.html",
        users=users,
        total_users=total_users,
        activos=activos,
        inactivos=inactivos,
        equipos=equipos,
        current_user=current_user
    )


# =========================
# CREAR USUARIO
# =========================
@app.route("/create", methods=["POST"])
@login_required
def create_user():

    accesos = ",".join(request.form.getlist("accesos"))

    new_user = User(
        nombre=request.form["nombre"],
        usuario=request.form["usuario"],
        correo=request.form["correo"],
        equipo=request.form["equipo"],
        jefe=request.form["jefe"],
        accesos=accesos,
        comentarios=request.form.get("comentarios", ""),
        activo=True
    )

    db.session.add(new_user)
    db.session.commit()

    # Notificación Teams si es invitado
    if current_user.role == "invitado":
        webhook = os.environ.get("TEAMS_WEBHOOK_URL")
        if webhook:
            requests.post(
                webhook,
                json={"text": f"Nuevo usuario agregado: {new_user.nombre}"}
            )

    return redirect(url_for("dashboard"))


# =========================
# EDITAR USUARIO
# =========================
@app.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_user(id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(id)

    if request.method == "POST":
        user.nombre = request.form["nombre"]
        user.usuario = request.form["usuario"]
        user.correo = request.form["correo"]
        user.equipo = request.form["equipo"]
        user.jefe = request.form["jefe"]
        user.accesos = ",".join(request.form.getlist("accesos"))
        user.comentarios = request.form.get("comentarios", "")
        user.activo = True if request.form.get("activo") == "true" else False

        db.session.commit()
        return redirect(url_for("dashboard"))

    return render_template("edit_user.html", user=user)


# =========================
# ELIMINAR USUARIO (ADMIN ONLY)
# =========================
@app.route("/delete_user/<int:id>", methods=["POST"])
@login_required
def delete_user(id):

    if current_user.role != "admin":
        return {"success": False}, 403

    user = User.query.get(id)

    if not user:
        return {"success": False}, 404

    try:
        # Eliminar vacaciones relacionadas primero
        Vacacion.query.filter_by(user_id=user.id).delete()

        deleted_data = {
            "success": True,
            "nombre": user.nombre,
            "usuario": user.usuario
        }

        db.session.delete(user)
        db.session.commit()

        return deleted_data

    except Exception as e:
        db.session.rollback()
        return {"success": False, "error": str(e)}, 500


# =========================
# EXPORTAR EXCEL
# =========================
@app.route("/export-excel")
@login_required
def export_excel():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    users = User.query.all()
    wb = Workbook()

    # HOJA 1
    ws = wb.active
    ws.title = "Usuarios"

    headers = [
        "ID", "Nombre", "Usuario", "Correo",
        "Equipo", "Jefe", "Accesos",
        "Comentarios", "Fecha Creación"
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78",
                              fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for user in users:
        ws.append([
            user.id,
            user.nombre,
            user.usuario,
            user.correo,
            user.equipo,
            user.jefe,
            user.accesos,
            user.comentarios,
            user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else ""
        ])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0
                     for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Reporte_Usuarios_Corporativo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================
# IMPORTAR EXCEL
# =========================
@app.route("/import_excel", methods=["POST"])
@login_required
def import_excel():

    file = request.files.get("file")

    if not file:
        flash("No se seleccionó archivo")
        return redirect(url_for("dashboard"))

    try:
        wb = load_workbook(file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            new_user = User(
                nombre=str(data.get("Nombre", "")).strip(),
                usuario=str(data.get("Usuario", "")).strip(),
                correo=str(data.get("Correo", "")).strip(),
                equipo=str(data.get("Equipo", "")).strip(),
                jefe=str(data.get("Jefe", "")).strip(),
                accesos=str(data.get("Accesos", "")).strip(),
                activo=True
            )

            db.session.add(new_user)
            count += 1

        db.session.commit()
        flash(f"{count} usuarios importados correctamente")

    except Exception as e:
        db.session.rollback()
        flash("Error al importar archivo")

    return redirect(url_for("dashboard"))


# =========================
# VACACIONES
# =========================
from datetime import date

@@app.route("/vacaciones")
@login_required
def vacaciones_view():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    hoy = date.today()

    # 🟢 Activas
    activas = Vacacion.query.filter(
        Vacacion.estado == "Aprobado",
        Vacacion.fecha_inicio <= hoy,
        Vacacion.fecha_fin >= hoy
    ).order_by(Vacacion.fecha_fin.asc()).all()

    # 🔜 Próximas (orden ascendente)
    proximas = Vacacion.query.filter(
        Vacacion.estado == "Aprobado",
        Vacacion.fecha_inicio > hoy
    ).order_by(Vacacion.fecha_inicio.asc()).all()

    # ⚪ Finalizadas
    finalizadas = Vacacion.query.filter(
        Vacacion.estado == "Aprobado",
        Vacacion.fecha_fin < hoy
    ).order_by(Vacacion.fecha_fin.desc()).all()

    # 🔢 Agregar días restantes dinámicamente
    for v in activas:
        v.dias_restantes = (v.fecha_fin - hoy).days

    for v in proximas:
        v.dias_para_iniciar = (v.fecha_inicio - hoy).days

    total_dias_anual = db.session.query(
        func.sum(Vacacion.dias_solicitados)
    ).filter(
        Vacacion.anio == hoy.year
    ).scalar() or 0

    vacaciones = Vacacion.query.order_by(
        Vacacion.fecha_inicio.desc()
    ).all()

    usuarios = User.query.filter_by(activo=True).all()

    return render_template(
        "vacaciones.html",
        vacaciones=vacaciones,
        activas=activas,
        proximas=proximas,
        finalizadas=finalizadas,
        usuarios=usuarios,
        total_dias_anual=total_dias_anual
    )
# =========================
# SOLICITAR VACACIONES
# =========================
@app.route("/solicitar_vacaciones", methods=["POST"])
@login_required
def solicitar_vacaciones():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    try:
        from datetime import datetime

        user_id = request.form["user_id"]
        user = User.query.get(user_id)

        fecha_inicio = datetime.strptime(
            request.form["fecha_inicio"], "%Y-%m-%d"
        ).date()

        fecha_fin = datetime.strptime(
            request.form["fecha_fin"], "%Y-%m-%d"
        ).date()

        if fecha_fin < fecha_inicio:
            flash("La fecha fin no puede ser menor que la fecha inicio")
            return redirect(url_for("vacaciones_view"))

        dias = (fecha_fin - fecha_inicio).days + 1

        # 🔴 Validar días disponibles
        if user.dias_vacaciones < dias:
            flash("El empleado no tiene suficientes días disponibles")
            return redirect(url_for("vacaciones_view"))

        # 🔵 DESCONTAR SOLO UNA VEZ
        user.dias_vacaciones -= dias

        nueva = Vacacion(
            user_id=user_id,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            dias_solicitados=dias,
            estado="Aprobado",
            registrado_por=current_user.username,
            anio=fecha_inicio.year
        )

        db.session.add(nueva)
        db.session.commit()

        flash("Vacaciones registradas correctamente")

    except Exception as e:
        db.session.rollback()
        flash("Error al registrar vacaciones")

    return redirect(url_for("vacaciones_view"))

# =========================
# ELIMINAR TODOS LOS USUARIOS (ADMIN ONLY)
# =========================


from models import db, User
from sqlalchemy import text

@app.route("/delete_all_users", methods=["POST"])
@login_required
def delete_all_users():

    if current_user.role != "admin":
        return {"success": False}, 403

    try:
        # Eliminar todos los registros
        db.session.query(User).delete()

        # Reiniciar autoincrement en SQLite
        db.session.execute(text("DELETE FROM sqlite_sequence WHERE name='user'"))

        db.session.commit()

        return {"success": True}

    except Exception as e:
        db.session.rollback()
        print("ERROR:", e)
        return {"success": False}, 500
    
# =========================
# ELIMINAR VACACIÓN (ADMIN ONLY)
# =========================

@app.route("/delete_vacacion/<int:id>", methods=["POST"])
@login_required
def delete_vacacion(id):

    if current_user.role != "admin":
        return {"success": False}, 403

    vacacion = Vacacion.query.get_or_404(id)
    user = User.query.get(vacacion.user_id)

    # 🔵 Devolver días al empleado
    user.dias_vacaciones += vacacion.dias_solicitados

    db.session.delete(vacacion)
    db.session.commit()

    return {"success": True}


# =========================
# EDITAR VACACIÓN (ADMIN ONLY)
# =========================


@app.route("/edit_vacacion/<int:id>", methods=["GET", "POST"])
@login_required
def edit_vacacion(id):

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    vacacion = Vacacion.query.get_or_404(id)

    from datetime import datetime

    if request.method == "POST":

        fecha_inicio = datetime.strptime(
            request.form["fecha_inicio"], "%Y-%m-%d"
        ).date()

        fecha_fin = datetime.strptime(
            request.form["fecha_fin"], "%Y-%m-%d"
        ).date()

        dias = (fecha_fin - fecha_inicio).days + 1

        vacacion.fecha_inicio = fecha_inicio
        vacacion.fecha_fin = fecha_fin
        vacacion.dias_solicitados = dias

        db.session.commit()

        return redirect(url_for("vacaciones_view"))

    return render_template("edit_vacacion.html", vacacion=vacacion)

# =========================
# EXPORTAR VACACIONES A EXCEL (ADMIN ONLY)
# =========================


@app.route("/exportar_vacaciones_excel")
@login_required
def exportar_vacaciones_excel():

    if current_user.role != "admin":
        return redirect(url_for("dashboard"))

    vacaciones = Vacacion.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vacaciones"

    headers = [
        "Empleado",
        "Fecha Inicio",
        "Fecha Fin",
        "Días",
        "Estado",
        "Registrado por",
        "Año"
    ]

    ws.append(headers)

    for v in vacaciones:
        ws.append([
            v.user.nombre,
            v.fecha_inicio,
            v.fecha_fin,
            v.dias_solicitados,
            v.estado,
            v.registrado_por,
            v.anio
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Reporte_Vacaciones.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )